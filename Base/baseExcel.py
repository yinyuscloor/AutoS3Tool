import csv
from typing import Union, Any
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl
"""
用于存放处理Excel和csv基础方法
"""

class BaseExcel(object):


    def open_sheet_Excel(self, filepath, sheet_name_or_index) -> Union[str, Worksheet]:
        '''打开表单-excel中某sheet'''
        worksheets = openpyxl.load_workbook(filepath)
        if isinstance(sheet_name_or_index, int): #若输入表单的序号
            worksheet = worksheets._sheets[sheet_name_or_index]
        elif isinstance(sheet_name_or_index, str): #若输入表单名
            worksheet = worksheets[sheet_name_or_index]
        else:
            return "sheet输入有误，请输入sheet_name_or_index"
        return worksheet


    def read_header_Excel(self,filepath, sheet_name_or_index):
        '''获取表单的表头'''
        worksheet = self.open_sheet_Excel(filepath,sheet_name_or_index)
        headers = []
        for i in worksheet[1]:
            headers.append(i.value)
        return headers


    def read_key_value_Excel(self,filepath,sheet_name_or_index):
        '''
        获取所有数据，且将表头中的内容与数据结合展示（以字典的形式）
        如：[{'序号':1,'会员卡号': '680021685898','机场名称':'上海机场'},{'序号':2,'会员卡号': '680021685899','机场名称':'广州机场'}]
        '''
        worksheet = self.open_sheet_Excel(filepath,sheet_name_or_index)
        rows = list(worksheet.rows)
        # 获取标题
        data = []
        for row in rows[1:]:
            rwo_data = []
            for cell in row:
                rwo_data.append(cell.value)
                # 列表转换成字典，与表头里的内容使用zip函数进行打包
            data_dict = dict(zip(self.read_header_Excel(filepath,sheet_name_or_index),rwo_data))
            data.append(data_dict)
        return data

    @staticmethod
    def write_change_Excel(filepath,sheet_name,row,column,data):
        '''写入Excel数据'''
        worksheets = openpyxl.load_workbook(filepath)
        worksheet = worksheets[sheet_name]
        # 修改单元格
        worksheet.cell(row,column).value = data
        # 保存
        worksheets.save(filepath)
        # 关闭
        worksheets.close()


    def csv_writer_csv(self,filepath):
        '''csv_writer'''
        #指定文件对象
        f = open(filepath,"w",encoding="utf-8",newline="")
        #基于稳健对象构建csv写入对象
        csv_writer = csv.writer(f)
        return csv_writer


if __name__ == '__main__':
    pass









